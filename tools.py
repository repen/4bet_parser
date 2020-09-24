import logging, os, csv
from openpyxl import Workbook, load_workbook

def save_csv(path, data):

    with open(path, "a", newline='') as f:
        writer = csv.writer(f, delimiter=',')
        writer.writerow(data)

def log(*args, **kwargs):
    name_logger = args[0]
    name_file   = args[1]
    write = kwargs.setdefault("write", False)
    # создаём logger
    logger = logging.getLogger(name_logger)
    logger.setLevel( logging.DEBUG )

    # создаём консольный handler и задаём уровень
    if not write:
        ch = logging.StreamHandler()
    else:
        # log write in disk
        ch = logging.FileHandler("/".join( [ os.getcwd(), name_file] ))

    ch.setLevel(logging.DEBUG)

    # создаём formatter
    formatter = logging.Formatter('%(asctime)s : line %(lineno)-3s : %(name)s : %(levelname)s : %(message)s')
    # %(lineno)d :
    # добавляем formatter в ch
    ch.setFormatter(formatter)

    # добавляем ch к logger
    logger.addHandler(ch)
    # Api
    # logger.debug('debug message')
    # logger.info('info message')
    # logger.warn('warn message')
    # logger.error('error message')
    # logger.critical('critical message')
    return logger



class ExelPlain:

    @staticmethod
    def load_xlsx(path):
        return load_workbook(path)

    def __init__(self):
        self.document = self.create_document()
        self.index = 0

    def create_document(self):
        wb = Workbook()
        # ws = wb.active
        return wb

    def create_sheet(self, title):
        ws = self.document.create_sheet(title, self.index)
        return self.document


def write(path, data):

    try:
        wb = ExelPlain.load_xlsx(path)
    except FileNotFoundError:
        wb = ExelPlain()
        wb = wb.create_sheet("Tennis")

    sheet_ranges = wb["Tennis"]
    sheet_ranges.append(data)
    wb.save(path)

if __name__ == "__main__":
    pass